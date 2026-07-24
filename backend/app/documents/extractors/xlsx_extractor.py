"""Formatting-preserving Excel cell extraction with per-cell structural context."""

from __future__ import annotations

from datetime import date, datetime, time
from dataclasses import replace
import io
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageSequence

from app.detection.types import StructuralContext
from app.documents.extractors.types import ExtractedDocument, LayoutToken
from app.documents.extractors.image_extractor import analyze_image_frames


class XlsxExtractionError(ValueError):
    pass


def _display_value(value: object) -> str:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def extract_xlsx(path: Path, *, tesseract_cmd: Path | None = None) -> ExtractedDocument:
    if tesseract_cmd is not None:
        import pytesseract
        pytesseract.pytesseract.tesseract_cmd = str(tesseract_cmd)
    parts: list[str] = []
    tokens: list[LayoutToken] = []
    contexts: list[StructuralContext] = []
    formula_count = 0
    forced = []
    embedded_images: list[dict[str, object]] = []
    try:
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    except Exception as exc:
        raise XlsxExtractionError("Excel workbook could not be opened") from exc
    try:
        for worksheet in workbook.worksheets:
            headers = {
                cell.column: _display_value(cell.value)
                for cell in worksheet[1]
                if cell.value is not None and cell.data_type != "f"
            }
            for row in worksheet.iter_rows():
                previous_label: str | None = None
                for cell in row:
                    if cell.value is None:
                        continue
                    if cell.data_type == "f":
                        formula_count += 1
                        continue
                    value = _display_value(cell.value)
                    if not value:
                        continue
                    if parts:
                        parts.append("\n")
                    start = sum(len(part) for part in parts)
                    parts.append(value)
                    end = start + len(value)
                    locator = f"{worksheet.title}|{cell.coordinate}"
                    for index, character in enumerate(value):
                        tokens.append(
                            LayoutToken(
                                text=character,
                                start=start + index,
                                end=start + index + 1,
                                locator=locator,
                                local_start=index,
                                local_end=index + 1,
                                source="xlsx_cell",
                            )
                        )
                    labels = []
                    if cell.row > 1 and cell.column in headers:
                        labels.append(headers[cell.column])
                    if previous_label:
                        labels.append(previous_label)
                    if labels:
                        contexts.append(StructuralContext(start, end, tuple(dict.fromkeys(labels))))
                    previous_label = value
            for image_index, drawing_image in enumerate(worksheet._images):
                locator = f"xlsx_image|{worksheet.title}|{image_index}"
                try:
                    image_bytes = drawing_image._data()
                    with Image.open(io.BytesIO(image_bytes)) as source_image:
                        frames = [frame.copy() for frame in ImageSequence.Iterator(source_image)]
                except Exception as exc:
                    raise XlsxExtractionError(
                        f"Embedded image {image_index + 1} on sheet {worksheet.title} could not be decoded"
                    ) from exc
                image_extracted = analyze_image_frames(frames, locator_prefix=locator)
                if parts:
                    parts.append("\n")
                offset = sum(len(part) for part in parts)
                parts.append(image_extracted.text)
                tokens.extend(
                    replace(token, start=token.start + offset, end=token.end + offset)
                    for token in image_extracted.tokens
                )
                forced.extend(
                    replace(candidate, start=candidate.start + offset, end=candidate.end + offset)
                    for candidate in image_extracted.preclassified_candidates
                )
                embedded_images.append(
                    {
                        "locator": locator,
                        "sheet_name": worksheet.title,
                        "image_index": image_index,
                        "start": offset,
                        "end": offset + len(image_extracted.text),
                        "extracted": image_extracted,
                    }
                )
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    return ExtractedDocument(
        text="".join(parts),
        tokens=tokens,
        structural_contexts=contexts,
        preclassified_candidates=forced,
        metadata={
            "sheet_count": len(sheet_names), "formula_count": formula_count,
            "embedded_image_count": len(embedded_images),
            "embedded_images": embedded_images,
        },
    )
